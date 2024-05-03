import logging
import re

import aiogram.filters
from openpyxl import load_workbook

from aiogram import types
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.enums.content_type import ContentType
from aiogram.filters import Command

from tgbot.loader import db, config
from tgbot.keyboards.reply import get_book_keyboard, get_test_by_book, get_passage_by_test, end_next_keyboard
from tgbot.keyboards.inline import get_books_inline
from tgbot.services.broadcaster import broadcast
from tgbot.misc.states import VocabularyTraining, NewWord

contest_router = Router()


@contest_router.message(F.text.casefold() == "finish")
async def cancel_handler(message: types.Message, state: FSMContext) -> None:
    """
    Allow user to cancel any action
    """
    current_state = await state.get_state()
    if current_state is None:
        return

    logging.info("Cancelling state %r", current_state)
    await state.clear()
    await message.answer(
        "Cancelled.",
        reply_markup=types.ReplyKeyboardRemove(),
    )


# training
@contest_router.message(Command('training'))
async def begin_registration(message: types.Message, state: FSMContext):
    markup = await get_book_keyboard()
    if len(markup.keyboard[0]) < 1:
        await message.answer("Kitob mavjud emas, <code>/new_book</code> - orqali yangi kitob qo'shing",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()
        return
    await message.answer("kitobni tanlang:", reply_markup=markup)
    await state.set_state(VocabularyTraining.BookTitle)


@contest_router.message(VocabularyTraining.BookTitle)
async def begin_registration(message: types.Message, state: FSMContext):
    markup = await get_test_by_book(message.text)
    if len(markup.keyboard[0]) < 1:
        await message.answer("test mavjud emas, <code>/new_test</code> - orqali yangi test raqamini qo'shing",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()
        return
    await message.answer("testni tanlang:", reply_markup=markup)
    await state.update_data({"book": message.text})
    await state.set_state(VocabularyTraining.TestNumber)


@contest_router.message(VocabularyTraining.TestNumber)
async def begin_registration(message: types.Message, state: FSMContext):
    data = await state.get_data()

    markup = await get_passage_by_test(int(message.text), data['book'])
    if len(markup.keyboard[0]) < 1:
        await message.answer("passage mavjud emas, <code>/new_passage</code> - orqali yangi passage qo'shing",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()
        return

    await message.answer("passage tanlang:", reply_markup=markup)
    await state.update_data({"test": int(message.text)})
    await state.set_state(VocabularyTraining.PassageNumber)


@contest_router.message(VocabularyTraining.PassageNumber)
async def begin_registration(message: types.Message, state: FSMContext):
    data = await state.get_data()
    words = await db.get_vocabulary_by_passage(int(message.text), int(data["test"]), data["book"], 15)
    if len(words) < 1:
        await message.answer(text="so'zlar mavjud emas. <code>/new_word</code> orqali yangi so'zlarni qo'shing",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()
        return
    index = 0
    word = words[index]["word"]
    await message.answer(text=f"{index+1}.Word: <b>{word}</b>\nDefinition: <i>{words[index]['definition']}</i>",
                         reply_markup=end_next_keyboard)
    question_count = len(words)
    answers = []
    for i in range(question_count):
        answers.append(words[i]["translation"])
    await state.update_data({"words": words, "question_count": question_count, "index": index+1,
                             "true_answers": answers, "answers": {}}
                            )
    await state.set_state(VocabularyTraining.Word)


@contest_router.message(VocabularyTraining.Word)
async def begin_registration(message: types.Message, state: FSMContext):
    data = await state.get_data()
    index = data["index"]
    words = data["words"]
    question_count = data["question_count"]
    answers: dict = data["answers"]
    answers[index-1] = message.text
    if index < question_count:
        word = words[index]["word"]
        text = f"{index+1}.Word: <b>{word}</b>\nDefinition: <i>{words[index]['definition']}</i>"
        await state.update_data({"index": index+1, "answers": answers})
        await message.answer(text=text)
        await state.set_state(VocabularyTraining.Word)
    else:
        answers = data["answers"]
        true_answers = data["true_answers"]
        keys = answers.keys()
        correct, wrong = 0, 0

        details = ""
        for key in keys:
            if true_answers[key] == answers[key]:
                mark = "✅"
                correct += 1
            else:
                mark = "❌"
                wrong += 1

            details += (f"{key + 1}. Correct: <code>{true_answers[key]}</code>, your_answer: <code>{answers[key]}</code>"
                        f" {mark}\n")

        await message.answer(f"<b>Quiz is over</b>\nCorrect:\t{correct}\nWrong:\t{wrong}\n\nDetails:\n{details}",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()


# new word
@contest_router.message(Command("new_word"))
async def new_word(message: types.Message, state: FSMContext):
    await message.answer("Kitob tanlang:", reply_markup=await get_book_keyboard())
    await state.set_state(NewWord.BookTitle)


@contest_router.message(NewWord.BookTitle)
async def new_word(message: types.Message, state: FSMContext):
    await state.update_data({"book": message.text})
    await message.answer("test raqamini tanlang:", reply_markup=await get_test_by_book(message.text))
    await state.set_state(NewWord.TestNumber)


@contest_router.message(NewWord.TestNumber)
async def new_word(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await message.answer("passageni tanlang:",
                         reply_markup=await get_passage_by_test(int(message.text), data["book"]))
    await state.update_data({"test": message.text})
    await state.set_state(NewWord.PassageNumber)


@contest_router.message(NewWord.PassageNumber)
async def new_word(message: types.Message, state: FSMContext):
    await state.update_data({"passage": message.text})
    await message.answer("so'zlar yozilgan excel faylni yuboring:", reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(NewWord.Word)


@contest_router.message(NewWord.Word)
async def new_word(message: types.Message, state: FSMContext):
    file = await message.bot.get_file(message.document.file_id)
    f = await message.bot.download_file(file.file_path)
    workbook = load_workbook(f)
    sheet = workbook.active

    await message.answer("processing...", reply_markup=types.ReplyKeyboardRemove())

    data = await state.get_data()

    ids = await db.return_ids_for_word_adding(data["book"], int(data["test"]), int(data["passage"]))
    row_data = {0: "", 1: "", 2: ""}
    for row in sheet.iter_rows():
        i = 0
        for cell in row:
            row_data[i] = cell.value
            i = i + 1
        try:
            await db.add_word(ids[0], ids[1], ids[2], row_data[0], row_data[1], row_data[2])
        except Exception as e:
            logging.warning(e)

    await message.answer(text="words added successfully", reply_markup=types.ReplyKeyboardRemove())
    await state.clear()


@contest_router.message(Command("inline"))
async def new_word(message: types.Message, state: FSMContext):
    await message.answer(text="inline keyboard", reply_markup=await get_books_inline())

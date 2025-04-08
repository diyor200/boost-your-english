import logging
import re
from unicodedata import category

import aiogram.filters
from aiogram.types import ReplyKeyboardRemove
from openpyxl import load_workbook

from aiogram import types
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.enums.content_type import ContentType
from aiogram.filters import Command

from tgbot.loader import db, config
from tgbot.keyboards.reply import (get_book_keyboard, get_passage_by_test, main_page_keyboard,
                                   finish_keyboard, get_collections, get_word_slices_markup, end_next_keyboard,
                                   go_menu_keyboard)
from tgbot.keyboards.inline import get_books_inline
from tgbot.services.broadcaster import broadcast
from tgbot.misc.states import Quiz, NewWord, CollectionWord, Collection

quiz_router = Router()


@quiz_router.message(F.text.casefold() == "finish")
async def cancel_handler(message: types.Message, state: FSMContext) -> None:
    """
    Allow user to cancel any action
    """
    current_state = await state.get_state()
    if current_state is None:
        return

    logging.info("Cancelling state %r", current_state)
    await state.clear()
    await message.answer(
        "Cancelled.",
        reply_markup=main_page_keyboard(),
    )


# training
@quiz_router.message(F.text == 'Quiz🤓')
async def begin_quiz(message: types.Message, state: FSMContext):
    markup = await get_book_keyboard()
    if len(markup.keyboard[0]) < 1:
        await message.answer("Bo'limlar mavjud emas, <code>/new_book</code> - orqali yangi kitob qo'shing",
                             reply_markup=main_page_keyboard())
        await state.clear()
        return
    await message.answer("bo'limni tanlang:", reply_markup=markup)
    await state.set_state(Quiz.Category)


@quiz_router.message(Quiz.Category)
async def choose_book(message: types.Message, state: FSMContext):
    c = await db.get_category_id_by_name(message.text)
    markup = await get_word_slices_markup(c['id'])
    if len(markup.keyboard[0]) < 1:
        await message.answer("test mavjud emas, <code>/new_test</code> - orqali yangi test raqamini qo'shing",
                             reply_markup=main_page_keyboard())
        await state.clear()
        return
    await message.answer("testni tanlang:", reply_markup=markup)
    await state.update_data({"category_id": c['id']})
    await state.set_state(Quiz.Slices)


@quiz_router.message(Quiz.Slices)
async def choose_passage(message: types.Message, state: FSMContext):
    if message.text == "🏠Bosh menu":
        await state.clear()
        await message.answer("Bo'limni tanlang", reply_markup=main_page_keyboard())
        return

    parts = message.text.split('-')
    if len(parts) < 2:
        await message.answer("Iltimos tugmalardan foydalaning:")
        await state.set_state(Quiz.Slices)
        return

    try:
        offset = int(parts[0])
        limit = int(parts[1]) - int(parts[0]) + 1
    except Exception as e:
        logging.error(e)
        await message.answer("Iltimos tugmalardan foydalaning:")
        await state.set_state(Quiz.Slices)
        return

    data = await state.get_data()
    words = await db.get_vocabularies_by_category_id_and_limits(category_id=data['category_id'], limit=limit,
                                                                offset=offset)
    if len(words) < 1:
        await message.answer(text="so'zlar mavjud emas. <code>/new_word</code> orqali yangi so'zlarni qo'shing",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()
        return

    index = 0
    word = words[index]["translation"]
    await message.answer(
        text=f"{index + 1}.Word: <b>{word}</b>",
        reply_markup=go_menu_keyboard())

    question_count = len(words)
    answers = []
    for i in range(question_count):
        answers.append(words[i]["word"])
    await state.update_data({"words": words, "question_count": question_count, "index": index + 1,
                             "true_answers": answers, "answers": {}}
                            )
    await state.set_state(Quiz.Word)


@quiz_router.message(Quiz.Word)
async def star_quiz(message: types.Message, state: FSMContext):
    data = await state.get_data()
    index = data["index"]
    words = data["words"]
    question_count = data["question_count"]
    answers: dict = data["answers"]
    answers[index-1] = message.text
    if index < question_count:
        word = words[index]["translation"]
        text = f"{index+1}.Word: <b>{word}</b>"
        await state.update_data({"index": index+1, "answers": answers})
        await message.answer(text=text)
        await state.set_state(Quiz.Word)
    else:
        answers = data["answers"]
        true_answers = data["true_answers"]
        keys = answers.keys()
        correct, wrong = 0, 0

        details = ""
        for key in keys:
            if true_answers[key] == answers[key]:
                mark = "✅"
                correct += 1
            else:
                mark = "❌"
                wrong += 1

            details += (f"{key + 1}. Correct: <code>{true_answers[key]}</code>, your_answer: <code>{answers[key]}</code>"
                        f" {mark}\n")

        await message.answer(f"<b>Quiz is over</b>\nCorrect:\t{correct}\nWrong:\t{wrong}\n\nDetails:\n{details}",
                             reply_markup=types.ReplyKeyboardRemove())
        await state.clear()


# new word
@quiz_router.message(Command("new_word"))
async def new_word(message: types.Message, state: FSMContext):
    await message.answer("bo'limni tanlang:", reply_markup=await get_book_keyboard())
    await state.set_state(NewWord.Category)


@quiz_router.message(NewWord.Category)
async def new_word(message: types.Message, state: FSMContext):
    category = await db.get_category_id_by_name(message.text)
    await state.update_data({"category_id": category['id']})
    await message.answer("fileni yuboring:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(NewWord.File)


@quiz_router.message(NewWord.File)
async def new_word(message: types.Message, state: FSMContext):
    file = await message.bot.get_file(message.document.file_id)
    f = await message.bot.download_file(file.file_path)
    workbook = load_workbook(f)
    sheet = workbook.active

    await message.answer("processing...", reply_markup=types.ReplyKeyboardRemove())


    data = await state.get_data()
    row_data = {0: "", 1: "", }
    for row in sheet.iter_rows():
        i = 0
        for cell in row:
            row_data[i] = cell.value
            i = i + 1
        try:
            await db.add_word_by_category_id(data['category_id'],row_data[0], row_data[1])
        except Exception as e:
            logging.warning(e)

    await message.answer(text="words added successfully", reply_markup=main_page_keyboard())
    await state.clear()

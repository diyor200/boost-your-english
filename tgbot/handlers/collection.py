import logging
import re

import aiogram.filters
from openpyxl import load_workbook

from aiogram import types
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.enums.content_type import ContentType
from aiogram.filters import Command

from tgbot.loader import db, config
from tgbot.keyboards.reply import main_page_keyboard, end_next_keyboard, get_collections
from tgbot.services.broadcaster import broadcast
from tgbot.misc.states import CollectionWord, Collection, CollectionWordTraining

collection_router = Router()


@collection_router.message(F.text.casefold() == "finish")
async def cancel_handler(message: types.Message, state: FSMContext) -> None:
    """
    Allow user to cancel any action
    """
    current_state = await state.get_state()
    if current_state is None:
        return

    logging.info("Cancelling state %r", current_state)
    await state.clear()
    await message.answer(
        "Cancelled.",
        reply_markup=main_page_keyboard(),
    )


@collection_router.message(F.text == 'Collection')
async def collection_handler(message: types.Message, state: FSMContext):
    try:
        collections = await db.select_all_collections(str(message.from_user.id))
        if len(collections) < 1:
            await message.answer(text="collectionlar mavjud emas, <code>/new_collection</code>"
                                      " orqali yangisini qo'shing")
            return

        await message.answer(text="collectionni tanlang", reply_markup=await get_collections(str(message.from_user.id)))
        await state.set_state(CollectionWordTraining.Collection)
    except Exception as e:
        logging.warning(e)
        await message.answer("muammo yuzaga keldi")


@collection_router.message(CollectionWordTraining.Collection)
async def get_collection(message: types.Message, state: FSMContext):
    try:
        collection_id = await db.get_collection_id(message.text, str(message.from_user.id))
        collection_words = await db.select_all_collection_words(collection_id[0])

        collection_words_count = len(collection_words)

        if collection_words_count < 1:
            await message.answer(text="so'zlar mavjud emasbuyruqlarni ko'rish uchun, /help ", reply_markup=main_page_keyboard())
            await state.clear()
            return

        index = 0
        await state.update_data({"collection_words": collection_words, "question_count": collection_words_count,
                                 "index": index+1})

        await message.answer(text=f"{index+1}.Word: <b>{collection_words[index]['word']}</b>"
                                  f"\nTranslation <code>{collection_words[index]['translation']}</code>",
                             reply_markup=end_next_keyboard)
        await state.set_state(CollectionWordTraining.Word)
    except Exception as e:
        logging.warning(e)
        await message.answer("muammo yuzaga keldi", reply_markup=main_page_keyboard())
        await state.clear()


@collection_router.message(CollectionWordTraining.Word)
async def begin_registration(message: types.Message, state: FSMContext):
    if message.text != "next":
        await message.answer("use keyboards, please")
        await state.set_state(CollectionWordTraining.Word)
        return

    data = await state.get_data()
    index = data["index"]
    words = data["collection_words"]
    question_count = data["question_count"]

    if index < question_count:
        text = (f"{index+1}.Word: <b>{words[index]['word']}</b>"
                f"\nTranslation: <code>{words[index]['translation']}</code>")

        await state.update_data({"index": index+1})
        await message.answer(text=text)
        await state.set_state(CollectionWordTraining.Word)
    else:
        await message.answer(f"<b>Train is over</b>. You can test your knowledge in Quiz section",
                             reply_markup=main_page_keyboard())
        await state.clear()



# new collection
@collection_router.message(Command("new_collection"))
async def new_collection(message: types.Message, state: FSMContext):
    await message.answer("Nomini kiriting:", reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(Collection.Title)


@collection_router.message(Collection.Title)
async def new_title(message: types.Message, state: FSMContext):
    try:
        await db.add_collection(message.text, str(message.from_user.id))
    except Exception as e:
        logging.warning(e)
        await message.answer("muammo yuzaga keldi")
    await state.clear()


@collection_router.message(Command("new_collection_word"))
async def new_collection(message: types.Message, state: FSMContext):
    await message.answer("Collectionni tanlang:", reply_markup=await get_collections(str(message.from_user.id)))
    await state.set_state(CollectionWord.Collection)


@collection_router.message(CollectionWord.Collection)
async def new_word(message: types.Message, state: FSMContext):
    await state.update_data({"collection": message.text})
    await message.answer("so'zlar yozilgan excel faylni yuboring:", reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(CollectionWord.Word)


@collection_router.message(CollectionWord.Word)
async def new_word(message: types.Message, state: FSMContext):
    file = await message.bot.get_file(message.document.file_id)
    f = await message.bot.download_file(file.file_path)
    workbook = load_workbook(f)
    sheet = workbook.active

    await message.answer("processing...", reply_markup=types.ReplyKeyboardRemove())

    data = await state.get_data()

    collection_id = await db.get_collection_id(data["collection"], str(message.from_user.id))
    row_data = {0: "", 1: ""}
    for row in sheet.iter_rows():
        i = 0
        for cell in row:
            row_data[i] = cell.value
            i = i + 1
        try:
            await db.add_collection_word(collection_id[0], row_data[0], row_data[1])
        except Exception as e:
            logging.warning(e)

    await message.answer(text="words added successfully", reply_markup=main_page_keyboard())
    await state.clear()
